# encoding=utf-8

import requests
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook, Workbook
import time

file_loc = 'F:/WeipuSpider/name_list.xlsx'
DOWNLOAD_URL = 'http://lib.cqvip.com/Search/SearchList'
header = {'Connection': 'keep-alive',
          'X-Requested-With': 'XMLHttpRequest',
          'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36',
          'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
          'Accept-Encoding': 'gzip, deflate',
          'Accept-Language': 'zh-CN,zh;q=0.9'}



def get_cookie(num):
    cookies = []
    cookies.append('_qddaz=QD.dvg5o9.qzku28.jtzlxy2x; user_behavior_flag=d27ac7e2-dc23-42f1-8e26-3c78c4ef900f; Hm_lvt_69fff6aaf37627a0e2ac81d849c2d313=1554198939,1554257648; __utma=164835757.174467248.1554262717.1554262717.1554262717.1; __utmz=164835757.1554262717.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); skybug=ea4aa775b9af539f55ca5843428f034b; ASP.NET_SessionId=2unchqhi4o3jaeddudc4y21f; search_isEnable=1')
    cookies.append('_qddaz=QD.dvg5o9.qzku28.jtzlxy2x; user_behavior_flag=d27ac7e2-dc23-42f1-8e26-3c78c4ef900f; Hm_lvt_69fff6aaf37627a0e2ac81d849c2d313=1554198939,1554257648; __utmt=1; __utmt_vip2=1; __utma=164835757.1762443098.1554257892.1554257892.1554257892.1; __utmc=164835757; __utmz=164835757.1554257892.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmb=164835757.1.10.1554257892; Hm_lpvt_69fff6aaf37627a0e2ac81d849c2d313=1554257892; ASP.NET_SessionId=jpspjxk3f2lg04yudf1xzpxo; skybug=c9d1538376cacc113f615e4718350312; search_isEnable=1')
    num = num % 2
    cookie_dict = {i.split("=")[0]:i.split("=")[-1] for i in cookies[num].split("; ")}



def get_name_list(file_loc):
    '''
    从excel中获取作者和作者对应的网址
    :param file_loc:excel文件地址
    :return:作者名
    '''
    workbook = load_workbook(file_loc)
    sheet1 = workbook.get_sheet_by_name('Sheet1')
    name = [i.value for i in sheet1['A']]
    return name




def get_web_info(name):
    url = DOWNLOAD_URL
    page_num = 1
    web_info = []
    cookie_dict = get_cookie(page_num)
    searchParamModel_1 = '''searchParamModel={"ObjectType":1,"SearchKeyList":[],"SearchExpression":null,"BeginYear":null,"EndYear":null,"UpdateTimeType":null,"JournalRange":null,"DomainRange":null,"ClusterFilter":"ZY=21#自动化与计算机技术","ClusterLimit":0,"ClusterUseType":"Article","UrlParam":"A='''
    searchParamModel_2 = '''","Sort":"0","SortField":null,"UserID":"0","PageNum":'''
    searchParamModel_3 = ''',"PageSize":'20',"SType":"","StrIds":"","IsRefOrBy":'0',"ShowRules":"  作者='''
    searchParamModel_4 = '''  ","IsNoteHistory":'0',"AdvShowTitle":null,"ObjectId":null,"ObjectSearchType":'0',"ChineseEnglishExtend":'0',"SynonymExtend":'0',"ShowTotalCount":'0',"AdvTabGuid":""}'''
    searchParamModel = searchParamModel_1 + name + searchParamModel_2 + str(page_num) + searchParamModel_3 + name + searchParamModel_4
    page = requests.post(url, headers=header, cookies=cookie_dict, data=searchParamModel.encode('utf-8'))
    soup = bs(page.content, 'lxml')
    # 获取总页面数
    try:
        total_page_num = int(soup.select_one(".layui-laypage-last").text)
    except:
        total_page_num = 1
    print('共计', total_page_num)
    print('当前', page_num)
    # 获取文章名、期刊名、关键词、作者信息
    paper_list = soup.find_all("dl")
    for i in paper_list:
        paper_title = i.select("a[index]")
        if paper_title:
            paper_title = paper_title[0].text
            author = ','.join([j.text for j in i.select("span[class='author'] a")])
            journal = i.select("span[class='from'] a")[0].text
            key_word = ','.join([j.text for j in i.select("span[class='subject'] a")])
            print(paper_title, author, journal, key_word)
            web_info.append([paper_title, author, journal, key_word])

    while page_num < total_page_num:
        page_num += 1
        cookie_dict = get_cookie(page_num)
        print('当前', page_num)
        time.sleep(5)
        searchParamModel = searchParamModel_1 + name + searchParamModel_2 + str(page_num) + searchParamModel_3 + name + searchParamModel_4
        page = requests.post(url, headers=header, cookies=cookie_dict, data=searchParamModel.encode('utf-8'))
        soup = bs(page.content, 'lxml')
        paper_list = soup.find_all("dl")
        for i in paper_list:
            paper_title = i.select("a[index]")
            if paper_title:
                paper_title = paper_title[0].text
                author = ','.join([j.text for j in i.select("span[class='author'] a")])
                journal = i.select("span[class='from'] a")[0].text
                key_word = ','.join([j.text for j in i.select("span[class='subject'] a")])
                print(paper_title, author, journal, key_word)
                web_info.append([paper_title, author, journal, key_word])
    return web_info

def main():
    name_list = get_name_list(file_loc)
    wb = Workbook()
    sheet = wb.active
    print(name_list)
    for name in name_list:
        print(name)
        info = get_web_info(name)
        for i in info:
            i += [name]
            sheet.append(i)
        time.sleep(5)
    wb.save(r'F:/WeipuSpider/paper_list.xlsx')

if __name__ == '__main__':
    main()